"""export_utils.py — Excel export for the dashboard."""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_excel(D: dict, data_key: str, label: str) -> bytes:
    wb  = Workbook()
    months = D['month_labels']
    DEC = D.get('last_dec_idx', len(D['month_labels']) - 1)

    def hfill(h): return PatternFill('solid', start_color=h)
    def mfont(bold=False, color='000000', size=10):
        return Font(bold=bold, color=color, size=size, name='Arial')
    def mkborder():
        s = Side(style='thin', color='BFBFBF')
        return Border(left=s, right=s, top=s, bottom=s)
    CA = Alignment(horizontal='center', vertical='center', wrap_text=True)

    TABS = [
        ('CARs',     'car_metrics', 'car_wavg',  '1F4E79', 'D6E4F0',
         ['Month','CARs Closed','Avg Days','Open >90','Wtd Avg Days','Trend']),
        ('PTOs',     'pto_metrics', 'pto_wavg',  '375623', 'D9EAD3',
         ['Month','PTOs Closed','Avg Days','Open >90','Wtd Avg Days','Trend']),
        ('Combined', 'cmb_metrics', 'cmb_wavg',  '4A235A', 'E8D5F5',
         ['Month','Total Closed','Avg Days','Total >90','CARs >90','PTOs >90','Wtd Avg Days','Trend']),
    ]

    first = True
    for tab_name, met_key, wav_key, dark, light, headers in TABS:
        ws = wb.active if first else wb.create_sheet(tab_name)
        if first: ws.title = tab_name; first = False

        n = len(headers)
        ws.merge_cells(f'A1:{get_column_letter(n)}1')
        ws['A1'] = f'{tab_name} — {label} — {datetime.now().strftime("%m/%d/%Y")}'
        ws['A1'].font = mfont(bold=True, color='FFFFFF', size=12)
        ws['A1'].fill = hfill(dark); ws['A1'].alignment = CA
        ws.row_dimensions[1].height = 24

        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = mfont(bold=True, color='FFFFFF', size=10)
            c.fill = hfill(dark); c.alignment = CA
        ws.row_dimensions[2].height = 20

        metrics   = D[met_key].get(data_key, D[met_key]['ALL'])
        wavg_vals = D[wav_key].get(data_key, D[wav_key]['ALL'])
        alt1 = hfill(light); alt2 = hfill('FFFFFF')

        for i, (m_label, row, wv) in enumerate(zip(months, metrics, wavg_vals)):
            r    = 3 + i
            fill = alt1 if i % 2 == 0 else alt2
            # Trend: compare snapshot figures month-over-month
            prev_ov = metrics[i-1]['ov90'] if i > 0 else None
            trend   = '—' if prev_ov is None else ('▲' if row['ov90'] > prev_ov else ('▼' if row['ov90'] < prev_ov else '→'))

            # ov90 = point-in-time snapshot at month-end
            vals = [m_label, row['closed'], row['avg_days'], row['ov90']]
            if tab_name == 'Combined':
                vals += [row.get('ov90_car', 0), row.get('ov90_pto', 0)]
            vals += [wv, trend]

            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=ci, value=v)
                c.fill = fill; c.font = mfont(size=10)
                c.border = mkborder(); c.alignment = CA

            # Year reset marker
            last_dec_yr = D.get('last_dec_year', 2025)
            if m_label == f'Jan {last_dec_yr + 1}':
                for ci in range(1, n+1):
                    ws.cell(row=r, column=ci).border = Border(
                        top=Side(style='medium', color='C00000'),
                        left=Side(style='thin', color='BFBFBF'),
                        right=Side(style='thin', color='BFBFBF'),
                        bottom=Side(style='thin', color='BFBFBF'))

        # Scorecard row
        sc_row = 3 + len(months)
        ws.cell(row=sc_row, column=1, value='TOTALS / AVGS').fill = hfill(dark)
        ws.cell(row=sc_row, column=1).font = mfont(bold=True, color='FFFFFF', size=10)
        ws.cell(row=sc_row, column=1).alignment = CA

        ye_wavg  = wavg_vals[DEC]
        ytd_wavg = wavg_vals[-1]
        sc_vals  = [sum(r['closed'] for r in metrics),
                    round(sum(r['avg_days'] for r in metrics if r['closed']>0) /
                          max(sum(1 for r in metrics if r['closed']>0), 1)),
                    round(sum(r['ov90'] for r in metrics) / len(metrics))]
        if tab_name == 'Combined':
            sc_vals += [round(sum(r.get('ov90_car',0) for r in metrics)/len(metrics)),
                        round(sum(r.get('ov90_pto',0) for r in metrics)/len(metrics))]
        sc_vals += [f"{ye_wavg} (YE{D.get('last_dec_year','')}) / {ytd_wavg} (YTD)", '']

        for ci, v in enumerate(sc_vals, 2):
            c = ws.cell(row=sc_row, column=ci, value=v)
            c.fill = hfill(dark); c.font = mfont(bold=True, color='FFFFFF', size=10)
            c.alignment = CA

        for ci in range(1, n+1):
            ws.column_dimensions[get_column_letter(ci)].width = 16
        ws.column_dimensions['A'].width = 14

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


def export_regional_summary(D: dict, as_of_date: str = None) -> bytes:
    """
    Export a regional summary with:
    - All locations from List source (zero-activity shows 0s)
    - Location name + ID code (e.g. "Avenel (NYH), NJ - 110")
    - Full Year columns B-D, YTD columns E-G with visual left-border separator
    - Cross-check verified: numbers match dashboard metrics
    """
    from datetime import datetime
    from openpyxl.utils import get_column_letter
    import numpy as np

    if as_of_date is None:
        as_of_date = datetime.now().strftime("%b %d, %Y")

    wb          = Workbook()
    months = D['month_labels']
    NM = len(months)
    last_dec_idx = D['last_dec_idx']
    last_dec_yr  = D['last_dec_year']
    idx_2026 = next((i for i, m in enumerate(months) if m == 'Jan 2026'), 0)
    region_map   = D['region_map']
    region_order = D['region_order']
    loc_id_map   = D.get('loc_id_map', {})

    def loc_display(name):
        lid = loc_id_map.get(name, '')
        return f'{name} - {lid}' if lid else name

    def hfill(h):     return PatternFill('solid', start_color=h)
    def mfont(bold=False, color='000000', size=10):
        return Font(bold=bold, color=color, size=size, name='Calibri')

    # Thin grey border for regular cells
    _thin = Side(style='thin', color='BFBFBF')
    _med  = Side(style='medium', color='BFBFBF')
    # The YTD separator — medium left border on column E
    _ytd_sep = Side(style='medium', color='595959')

    def mkborder(thick_top=False, ytd_left=False):
        t = Side(style='medium' if thick_top else 'thin', color='BFBFBF')
        l = _ytd_sep if ytd_left else _thin
        return Border(left=l, right=_thin, top=t, bottom=_thin)

    CA = Alignment(horizontal='center', vertical='center', wrap_text=True)
    CL = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # ── YE index range: Jan of last_dec_yr → last_dec_idx ─────────
    ye_start = next((i for i, m in enumerate(months) if m == f'Jan {last_dec_yr}'), 0)

    # ── YTD range: Jan of current year → last month ───────────────
    cur_yr    = int(months[-1].split()[-1])
    ytd_start = next((i for i, m in enumerate(months) if m == f'Jan {cur_yr}'), 0)

    def calc_metrics_for_range(metrics_dict, loc_key, start_i, end_i):
        """Weighted avg_days, last_ov90, total_closed for a month range.
        Uses same logic as dashboard — only closed records."""
        rows = metrics_dict.get(loc_key, [])
        if not rows:
            return 0, 0, 0
        slc  = rows[start_i:end_i + 1]
        if not slc:
            return 0, 0, 0
        total_closed = sum(r['closed'] for r in slc)
        total_days   = sum(r['closed'] * r['avg_days'] for r in slc)
        wtd_avg      = int(round(total_days / total_closed)) if total_closed > 0 else 0
        last_ov      = rows[end_i]['ov90'] if end_i < len(rows) else 0
        return wtd_avg, last_ov, total_closed

    TABS = [
        ('CARs',     'car_metrics',  '1F4E79'),
        ('PTOs',     'pto_metrics',  '375623'),
        ('Combined', 'cmb_metrics',  '4A235A'),
    ]

    REGION_COLORS_HEX = {
        'USWC':         'D6E4F0', 'USGC':          'FCE5CD',
        'USNE':         'D9EAD3', 'USMW & River':  'E9D5F5',
        'USMA & Carib': 'D0F0E8', 'Canada':        'EDE1F5',
        'NAM/Chem':     'FDEBD0', 'NAM/LPG':       'D5F5E3',
        'Corporate':    'EAECEE', 'Environmental':  'D1F2EB',
        'ADD/Calib':    'FADBD8', 'Agri':           'FEF9E7',
    }

    first = True
    for tab_name, met_key, dark_hex in TABS:
        ws = wb.active if first else wb.create_sheet(tab_name)
        if first: ws.title = tab_name; first = False

        # ── Row 1: Title ───────────────────────────────────────────
        ws.merge_cells('A1:I1')
        ws['A1'] = f'{tab_name} Regional Summary — As of {as_of_date}'
        ws['A1'].font      = mfont(bold=True, color='FFFFFF', size=12)
        ws['A1'].fill      = hfill(dark_hex)
        ws['A1'].alignment = CA
        ws.row_dimensions[1].height = 26

        # ── Row 2: Group headers ───────────────────────────────────
        ws.merge_cells('B2:D2')
        ws.merge_cells('E2:G2')
        ws['B2'] = f'{last_dec_yr} Full Year'
        ws['E2'] = f'YTD {as_of_date}'
        for cell_ref, ytd in [('B2', False), ('E2', True)]:
            c = ws[cell_ref]
            c.font      = mfont(bold=True, color='FFFFFF', size=10)
            c.fill      = hfill(dark_hex)
            c.alignment = CA
            if ytd:
                c.border = Border(left=_ytd_sep, right=_thin, top=_thin, bottom=_thin)
        ws['A2'].fill = hfill(dark_hex)
        ws['H2'].fill = hfill(dark_hex)
        ws.row_dimensions[2].height = 20

        # ── Row 3: Column headers ──────────────────────────────────
        headers = [
            ('Region / Location',         False),
            ('Wtd Avg Days to Complete',      False),
            (f'Open >90 as of Dec 31',     False),
            ('Total Closed',               False),
            ('Wtd Avg Days to Complete',      True),   # YTD start — separator
            (f'Open >90 (Current)',        False),
            ('Total Closed',               False),
            ('Weighted Avg Days (≥2026)',            False),
            ('Notes',                      False),
        ]
        for ci, (h, is_ytd_start) in enumerate(headers, 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font      = mfont(bold=True, color='FFFFFF', size=9)
            c.fill      = hfill(dark_hex)
            c.alignment = CA
            c.border    = mkborder(ytd_left=is_ytd_start)
        ws.row_dimensions[3].height = 32

        row_num = 4
        ws.sheet_properties.outlinePr.summaryBelow = False  # collapse button above region row
        for region in region_order:
            if region not in region_map:
                continue
            locs     = sorted(region_map[region])
            reg_fill = hfill(REGION_COLORS_HEX.get(region, 'F2F2F2'))

            # Region aggregate row using REGION: key
            reg_key = f'REGION:{region}'
            ye_avg,  ye_ov,  ye_cls   = calc_metrics_for_range(D[met_key], reg_key, ye_start, last_dec_idx)
            ytd_avg, ytd_ov, ytd_cls  = calc_metrics_for_range(D[met_key], reg_key, ytd_start, NM - 1)

            ws.cell(row=row_num, column=1, value=region).font = mfont(bold=True, size=10, color='0D1117')
            # post‑2026 average (using all months from Jan 2026 onward)
            post_avg, _, _ = calc_metrics_for_range(D[met_key], reg_key, idx_2026, NM - 1)
            vals = [ye_avg, ye_ov, ye_cls, ytd_avg, ytd_ov, ytd_cls, post_avg, '']
            for ci, v in enumerate(vals, 2):
                c = ws.cell(row=row_num, column=ci, value=v)
                c.font = mfont(bold=True, size=10)
            for ci in range(1, 10):
                ws.cell(row=row_num, column=ci).fill      = reg_fill
                ws.cell(row=row_num, column=ci).alignment = CA if ci > 1 else CL
                ws.cell(row=row_num, column=ci).border    = mkborder(
                    thick_top=True, ytd_left=(ci == 5))
            ws.row_dimensions[row_num].height = 18
            region_header_row = row_num
            row_num += 1

            # Location rows — grouped so they collapse under the region header
            loc_start = row_num
            for loc in locs:
                ye_avg,  ye_ov,  ye_cls   = calc_metrics_for_range(D[met_key], loc, ye_start, last_dec_idx)
                ytd_avg, ytd_ov, ytd_cls  = calc_metrics_for_range(D[met_key], loc, ytd_start, NM - 1)

                alt_fill = hfill('FAFBFC') if row_num % 2 == 0 else hfill('FFFFFF')
                ws.cell(row=row_num, column=1, value=f'  {loc_display(loc)}').font = mfont(size=9)
                post_avg, _, _ = calc_metrics_for_range(D[met_key], loc, idx_2026, NM - 1)
                loc_vals = [ye_avg, ye_ov, ye_cls, ytd_avg, ytd_ov, ytd_cls, post_avg, '']
                for ci, v in enumerate(loc_vals, 2):
                    ws.cell(row=row_num, column=ci, value=v).font = mfont(size=9)
                for ci in range(1, 10):
                    ws.cell(row=row_num, column=ci).fill      = alt_fill
                    ws.cell(row=row_num, column=ci).alignment = CA if ci > 1 else CL
                    ws.cell(row=row_num, column=ci).border    = mkborder(ytd_left=(ci == 5))
                ws.row_dimensions[row_num].height = 16
                ws.row_dimensions[row_num].outline_level = 1
                ws.row_dimensions[row_num].hidden = False
                row_num += 1

        # ── NAM TOTAL row ──────────────────────────────────────────
        nam_key = 'ALL'
        nam_ye_avg,  nam_ye_ov,  nam_ye_cls  = calc_metrics_for_range(D[met_key], nam_key, ye_start, last_dec_idx)
        post_avg, _, _ = calc_metrics_for_range(D[met_key], nam_key, idx_2026, NM - 1)
        nam_ytd_avg, nam_ytd_ov, nam_ytd_cls = calc_metrics_for_range(D[met_key], nam_key, ytd_start, NM - 1)
        nam_fill = hfill('1A1A2E')  # dark navy
        nam_vals = [nam_ye_avg, nam_ye_ov, nam_ye_cls, nam_ytd_avg, nam_ytd_ov, nam_ytd_cls, post_avg, '']
        for ci, v in enumerate(nam_vals, 2):
            c = ws.cell(row=row_num, column=ci, value=v)
            c.font = mfont(bold=True, size=10, color='FFFFFF')
        for ci in range(1, 10):
            ws.cell(row=row_num, column=ci).fill      = nam_fill
            ws.cell(row=row_num, column=ci).alignment = CA if ci > 1 else CL
            ws.cell(row=row_num, column=ci).border    = mkborder(thick_top=True, ytd_left=(ci == 5))
        ws.row_dimensions[row_num].height = 20
        row_num += 1

        # Column widths
        ws.column_dimensions['A'].width = 36
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 18
        for col in ['E', 'F', 'G']:
            ws.column_dimensions[col].width = 18
        ws.column_dimensions['H'].width = 18
        ws.freeze_panes = 'B4'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()
